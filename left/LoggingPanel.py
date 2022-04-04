import threading
from tkinter import*
from tkinter import ttk, messagebox, filedialog
from turtle import left
from openpyxl import Workbook, load_workbook
import datetime


class LoggingPanel(LabelFrame):

    column1_x = 5
    row1_y = 3
    row2_y = 30
    row3_y = 66

    relHeight = 0.29

    def __init__(self, master, controller):
        super().__init__(master)
        self.backgroundColor = master.backgroundColor
        self.config(labelwidget=ttk.Label(
            text="LoggingPanel", font=controller.labelFrameFont, foreground="grey",
            background=self.backgroundColor), background=self.backgroundColor)
        self.controller = controller

        btnStyle = ttk.Style()
        btnStyle.configure('my.TButton', font=controller.buttonFont)

        self.filename = ""
        self.printNote = True

        self.fileNameLabel = Label(self, text="file", font=controller.labelFont,
                                   background=self.backgroundColor, foreground="black", padx=10, pady=5, justify="left")
        self.fileNameLabel.place(relwidth=0.93, relheight=self.relHeight-0.08,
                                 x=self.column1_x, y=self.row1_y)

        self.noteLabel = Label(self, text="Note:", font=controller.labelFont,
                               background=self.backgroundColor, foreground="black", border=2, relief="groove", padx=10, pady=5)
        self.noteLabel.place(relwidth=0.25, relheight=self.relHeight,
                             x=self.column1_x, y=self.row2_y)

        self.note = Entry(self, font=('Arial', 15),
                          borderwidth=0, highlightthickness=2, highlightbackground="grey", highlightcolor="grey")
        self.note.place(relwidth=0.65, relheight=self.relHeight,
                        x=82, y=self.row2_y)

        self.openBtn = ttk.Button(
            self, text="open", command=self.openExcelFile, style="my.TButton")
        self.openBtn.place(relwidth=0.3, relheight=self.relHeight,
                           x=self.column1_x, y=self.row3_y)

        self.exportBtn = ttk.Button(
            self, text="export", command=self.export2xlsx, style="my.TButton")
        self.exportBtn.place(
            relwidth=0.3, relheight=self.relHeight, x=100, y=self.row3_y)

    def updateFileLabel(self):
        filePath = self.filename.split("/")
        self.fileNameLabel.config(text="file: "+filePath.pop())

    def export2xlsx(self):
        try:
            wb = load_workbook(self.filename)
        except:
            wb = Workbook()
            wb.create_sheet("all")
            wb.create_sheet("once")
            self.filename = "./excel/"+datetime.datetime.now().strftime("%d-%m-%y %H:%M")+".xlsx"
            self.updateFileLabel()
        try:
            ws = wb["all"]
        except:
            wb.create_sheet("all")
            ws = wb["all"]
        ws.append(["note", self.note.get()])
        ws.append(self.controller.dataName)

        for i in range(len(self.controller.data[0])-1):
            ws.append(self.controller.data[j][i+1]
                      for j in range(len(self.controller.data)))

        ws.append([])
        wb.save(self.filename)
        print("saved " + self.filename + ".xlsx")

    def openExcelFile(self):
        self.filename = filedialog.askopenfilename(
            initialdir="./excel", title="Select a excel file", filetypes=(("all", "."), ("xlsx", "*.xlsx")))
        self.updateFileLabel()

    def autoExport(self, cmd):
        packetArray = self.controller.packetString.strip("\n").split("-")
        try:
            wb = load_workbook(self.filename)
        except:
            wb = Workbook()
            wb.create_sheet("all")
            wb.create_sheet("once")
            self.filename = "./excel/"+datetime.datetime.now().strftime("%d-%m-%y %H:%M")+".xlsx"
            self.updateFileLabel()
        try:
            ws = wb["once"]
        except:
            wb.create_sheet("once")
        ws = wb["once"]
        if self.printNote:
            ws.append(["note", self.note.get()])
        if cmd == "save":
            packetArray.pop(0)
            print(packetArray)
            ws.append(packetArray)
            wb.save(self.filename)
            self.printNote = False
        if cmd == "end":
            ws.append([])
            wb.save(self.filename)
            self.printNote = True
            print("saved " + self.filename + ".xlsx")
