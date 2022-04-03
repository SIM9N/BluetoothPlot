import datetime
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from tkinter import ttk, messagebox
from tkinter import*
import serial
import serial.tools.list_ports
import matplotlib

matplotlib.use("TkAgg")


labelFont = ("Courier New", "11")
dropDownFont = ("Courier New", "12")

root = Tk()
root.title("BluetoothPlot")
root.config(bg="grey")
root.geometry("1000x600")
root.minsize(1000, 600)
style = ttk.Style()
style.theme_use("default")

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=3)

leftWrapperFrame = Frame(root, bg="#8ab2b4", bd=2, relief=RIDGE)
leftWrapperFrame.grid(column=0, row=0, sticky="nesw")

rightWrapperFrame = Frame(root, bg="#c9eeeb", bd=2, relief=RIDGE)
rightWrapperFrame.grid(column=1, row=0, sticky="nesw")

# a function for ttkComboBox


def defocus(event):
    event.widget.master.focus_set()


# --- Start Left Part ---#
leftWrapperFrame.grid_columnconfigure(0, weight=1)
leftWrapperFrame.grid_rowconfigure(0, weight=1)
leftWrapperFrame.grid_rowconfigure(1, weight=3)

# serial port
serialFrame = Frame(leftWrapperFrame, bg="#a4b2b0")
serialFrame.grid(column=0, row=0, padx=5, pady=5, sticky="nesw")

# selecting the serial port
serialPortOptions = ["0"]
avaPorts = serial.tools.list_ports.comports()
serialObj = serial.Serial()

# append all the available ports into srialPortOptions
for onePort in avaPorts:
    serialPortOptions.append(str(onePort).split(" ")[0])

portLabel = Label(serialFrame, text="Port:", font=labelFont)

dropDownSerialOptions = ttk.Combobox(
    serialFrame,
    value=serialPortOptions,
    font=dropDownFont,
    state="readonly",
)

dropDownSerialOptions.current(0)
dropDownSerialOptions.bind("<FocusIn>", defocus)

# selecting the serial baudrate
baudrateOptions = [1800, 2400, 4800, 9600,
                   19200, 28800, 38400, 57600, 76800, 115200]

baudrateLabel = Label(serialFrame, text="Baudrate:", font=labelFont)

dropDownBaudrateOptions = ttk.Combobox(
    serialFrame,
    value=baudrateOptions,
    font=dropDownFont,
    state="readonly",
)
dropDownBaudrateOptions.current(len(baudrateOptions) - 1)
dropDownBaudrateOptions.bind("<FocusIn>", defocus)

# connect the selected comport


def initComPort():
    try:
        serialObj.port = dropDownSerialOptions.get()
        serialObj.baudrate = dropDownBaudrateOptions.get()
        serialObj.close()
        serialObj.open()
    except IOError as e:
        messagebox.showerror(
            "Error",
            "Error, can't connect to "
            + dropDownSerialOptions.get()
            + "\n\n"
            + "( "
            + str(e)
            + " )",
        )


connectBtn = Button(serialFrame, text="connect",
                    command=initComPort, fg='black', borderwidth=0)

disconnectBtn = Button(
    serialFrame, text="disconnect", command=lambda: serialObj.close(), borderwidth=0
)

# Bluetooth data print
dataFrame = Frame(leftWrapperFrame, bg="#bbc1c8")
dataFrame.grid(column=0, row=1, padx=5, pady=5, sticky="nesw")

textData = Text(dataFrame, wrap=NONE, font=labelFont,
                takefocus=0, bg="white", fg="black")

dataCanvasScrollbar = ttk.Scrollbar(
    dataFrame, orient=VERTICAL, command=textData.yview)
dataCanvasScrollbar2 = ttk.Scrollbar(
    dataFrame, orient=HORIZONTAL, command=textData.xview
)

textData["yscroll"] = dataCanvasScrollbar.set
textData["xscroll"] = dataCanvasScrollbar2.set

varName = [""]
startCollectingData = False
Data = []
recentPacketString = ""


def printSerialPortData():
    if serialObj.isOpen() and serialObj.in_waiting:
        recentPacket = serialObj.readline()
        global recentPacketString
        try:
            recentPacketString = recentPacket.decode("utf")
        except:
            print("skip")
        textData.insert("end", recentPacketString)
        textData.see(END)

# format of the start cmd "bp-dataTitle1-dataTitle2-\n"
# format of the stop cmd "stop\n"


def checkStartCMD():
    global startCollectingData
    packetArray = recentPacketString.strip("\n").split("-")
    if packetArray[0] == "dataName":
        print("found dataName")
        global varName
        varName = packetArray
        varName.pop(0)
        global Data
        Data = [[None] for i in range(len(varName))]
        print(Data)
        print(len(Data))
        print(len(varName))
        dropDownX.config(values=varName)
        dropDownY.config(values=varName)
        startCollectingData = True
    elif packetArray[0] == "stop":
        print("stop")
        startCollectingData = False

# append the bluetooth data into Data[]


def appendData():
    if startCollectingData:
        packetArray = recentPacketString.strip("\n").split(",")
        # print(packetArray)
        if len(packetArray) != 1:
            try:
                for i in range(len(varName)):
                    Data[i].append(packetArray[i])
            except:
                print("append failed")


# clean the text in dataFrame
def cleanData():
    textData.delete(1.0, END)


cleanBtn = Button(serialFrame, text="clean", command=cleanData, borderwidth=0)


def export2xlsx():
    currentTime = datetime.datetime.now()
    wb = Workbook()
    ws = wb.active
    ws.append(varName)
    try:
        for i in range(len(Data[0])-1):
            ws.append(Data[j][i+1] for j in range(len(Data)))
    except IndexError as ex:
        print("IndexError when exporting data")

    wb.save("./excel/"+currentTime.strftime("%d-%m-%y %H:%M:%S")+".xlsx")


exportBtn = Button(serialFrame, text="export",
                   command=export2xlsx, borderwidth=0)

wb_auto = Workbook()
ws_auto = wb_auto.active


def autoExport():
    currentTime = datetime.datetime.now()
    packetArray = recentPacketString.strip("\n").split("-")
    if packetArray[0] == "ae":
        print("found ae")
        packetArray.pop(0)
        print(packetArray)
        ws_auto.append(packetArray)
    if packetArray[0] == "end":
        print("found end")
        ws_auto.append([])
        wb_auto.save("./excel/"+currentTime.strftime("%d-%m-%y")+"_auto.xlsx")

# --- End Left Part ---#


# --- Start Right Part ---#
fig = Figure(figsize=(0.1, 0.1), dpi=100)
ax1 = fig.add_subplot(221)
ax2 = fig.add_subplot(222)
ax3 = fig.add_subplot(223)
ax4 = fig.add_subplot(224)


class Toolbar(NavigationToolbar2Tk):
    def set_message(self, s):
        pass


graphCanvas = FigureCanvasTkAgg(fig, rightWrapperFrame)
graphCanvas.draw()
graphToolBar = Toolbar(graphCanvas, rightWrapperFrame, pack_toolbar=False)
graphToolBar.update()

controlPanelFrame = Frame(rightWrapperFrame, bg="#d7efd2")

dropDownGraph = ttk.Combobox(
    controlPanelFrame,
    value=["graph1", "graph2", "graph3", "graph4"],
    font=dropDownFont,
    state="readonly",
    width=7,
)
graphLabel = Label(
    controlPanelFrame, text="plot", bg="#d7efd2", font=labelFont, fg="black"
)
graphLabel.pack(side=LEFT, padx=5, pady=5, fill=Y)

dropDownGraph.current(0)
dropDownGraph.bind("<FocusIn>", defocus)
dropDownGraph.pack(side=LEFT, padx=5, pady=7, fill=Y)

dropDownX = ttk.Combobox(
    controlPanelFrame,
    value=varName,
    font=dropDownFont,
    state="readonly",
    width=10,
)
xLabel = Label(controlPanelFrame, text="X",
               bg="#d7efd2", font=labelFont, fg="black")
xLabel.pack(side=LEFT, padx=5, pady=5, fill=Y)

dropDownX.current(0)
dropDownX.bind("<FocusIn>", defocus)
dropDownX.pack(side=LEFT, padx=5, pady=7, fill=Y)

dropDownY = ttk.Combobox(
    controlPanelFrame,
    value=varName,
    font=dropDownFont,
    state="readonly",
    width=10,
)

yLabel = Label(controlPanelFrame, text="Y",
               bg="#d7efd2", font=labelFont, fg="black")
yLabel.pack(side=LEFT, padx=5, pady=5, fill=Y)

dropDownY.current(0)
dropDownY.bind("<FocusIn>", defocus)
dropDownY.pack(side=LEFT, padx=5, pady=7, fill=Y)


def plotIt():
    graphNum = dropDownGraph.get()
    xValue = Data[dropDownX.current()]
    yValue = Data[dropDownY.current()]

    if graphNum == "graph1":
        ax1.cla()
        ax1.plot(
            xValue,
            yValue,
            color="#444444",
            linestyle="--",
            label="label",
        )

        ax1.set_title(dropDownY.get() + " - " + dropDownX.get())
        ax1.set_xlabel(dropDownX.get())
        ax1.set_ylabel(dropDownY.get())

    if graphNum == "graph2":
        ax2.cla()
        ax2.plot(xValue, yValue)
        ax2.set_title(dropDownY.get() + " - " + dropDownX.get())
        ax2.set_xlabel(dropDownX.get())
        ax2.set_ylabel(dropDownY.get())

    if graphNum == "graph3":
        ax3.cla()
        ax3.plot(xValue, yValue)
        ax3.set_title(dropDownY.get() + " - " + dropDownX.get())
        ax3.set_xlabel(dropDownX.get())
        ax3.set_ylabel(dropDownY.get())

    if graphNum == "graph4":
        ax4.cla()
        ax4.plot(xValue, yValue)
        ax4.set_title(dropDownY.get() + " - " + dropDownX.get())
        ax4.set_xlabel(dropDownX.get())
        ax4.set_ylabel(dropDownY.get())
    fig.tight_layout()
    controlPanelFrame.update()
    graphCanvas.draw()


def customPlotIt():
    plt.plot(
        Data[dropDownX.current()],
        Data[dropDownY.current()],
        color="#444444",
        linestyle="--",
        label="label",
    )
    plt.title("cutomGraph")
    plt.xlabel(dropDownX.get())
    plt.ylabel(dropDownY.get())
    plt.show()


plotBtn = Button(
    controlPanelFrame, text="plot",  command=plotIt, borderwidth=0, highlightthickness=0
)
plotBtn.pack(side=LEFT, padx=5, pady=5, fill=Y)

customPlotBtn = Button(
    controlPanelFrame, text="custom",  command=customPlotIt, borderwidth=0, highlightthickness=0
)
customPlotBtn.pack(side=LEFT, padx=5, pady=5, fill=Y)


# --- End Right Part ---#


def place():

    portLabel.place(relheight=0.2, relwidth=1 - 0.8, x=2, y=0)

    dropDownSerialOptions.place(
        anchor=NE, relwidth=0.75, relheight=0.2, x=serialFrame.winfo_width() - 2, y=0
    )

    baudrateLabel.place(
        relheight=0.2, relwidth=1 - 0.7, x=2, y=10 + portLabel.winfo_height()
    )

    dropDownBaudrateOptions.place(
        anchor=NE,
        relwidth=0.65,
        relheight=0.2,
        x=serialFrame.winfo_width() - 2,
        y=10 + dropDownSerialOptions.winfo_height(),
    )

    connectBtn.place(
        relwidth=0.47,
        relheight=0.2,
        x=2,
        y=20 + baudrateLabel.winfo_height() + portLabel.winfo_height(),
    )

    disconnectBtn.place(
        anchor=NE,
        relwidth=0.47,
        relheight=0.2,
        x=serialFrame.winfo_width() - 2,
        y=20
        + dropDownSerialOptions.winfo_height()
        + dropDownBaudrateOptions.winfo_height(),
    )

    cleanBtn.place(
        relwidth=0.47,
        relheight=0.2,
        x=2,
        y=25
        + baudrateLabel.winfo_height()
        + portLabel.winfo_height()
        + connectBtn.winfo_height(),
    )

    exportBtn.place(
        anchor=NE,
        relwidth=0.47,
        relheight=0.2,
        x=serialFrame.winfo_width() - 2,
        y=25
        + baudrateLabel.winfo_height()
        + portLabel.winfo_height()
        + connectBtn.winfo_height(),
    )

    textData.place(
        anchor=NW,
        x=0,
        y=0,
        relheight=0.95,
        relwidth=0.92,
    )

    dataCanvasScrollbar.place(
        anchor=NE,
        x=serialFrame.winfo_width(),
        y=0,
        relwidth=0.08,
        relheight=1,
    )

    dataCanvasScrollbar2.place(
        x=0,
        y=textData.winfo_height(),
        relheight=0.05,
        relwidth=0.92,
    )

    graphToolBar.place(
        anchor=SW,
        x=0,
        y=rightWrapperFrame.winfo_height() - 5,
        relwidth=1,
    )

    graphCanvas.get_tk_widget().place(
        anchor=NW,
        x=0,
        y=controlPanelFrame.winfo_height(),
        height=rightWrapperFrame.winfo_height()
        - graphToolBar.winfo_height()
        - controlPanelFrame.winfo_height(),
        relwidth=1,
    )

    controlPanelFrame.place(
        anchor=NW,
        x=0,
        y=0,
        relwidth=1,
        height=40,
    )


while True:
    root.update_idletasks()
    root.update()
    printSerialPortData()
    checkStartCMD()
    appendData()
    autoExport()
    place()
