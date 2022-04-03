from tkinter import*
from tkinter import ttk, messagebox, filedialog
from turtle import left
from openpyxl import Workbook, load_workbook
import datetime

LabelFrameFont = ("Courier New", "14", "bold")
ButtonFont = ("Courier New", "11")
labelFont = ("Courier New", "14")


class LoggingPanel(LabelFrame):
    def __init__(self, master):
        super().__init__(master)
        self.backgroundColor = master.backgroundColor
        self.config(labelwidget=ttk.Label(
            text="LoggingPanel", font=LabelFrameFont, foreground="grey",
            background=self.backgroundColor), background=self.backgroundColor)

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=1)
        self.columnconfigure(3, weight=1)
        self.columnconfigure(4, weight=1)
        self.columnconfigure(5, weight=1)

        btnStyle = ttk.Style()
        btnStyle.configure('my.TButton', font=ButtonFont)

        self.filename = ""

        self.noteLabel = Label(self, text="Note:", font=labelFont,
                               background=self.backgroundColor, foreground="black", border=2, relief="groove", padx=10, pady=5)
        self.noteLabel.grid(column=0, row=0, columnspan=2,
                            padx=(4, 0), pady=0, sticky="WE")

        self.note = Entry(self, font=('Arial', 20),
                          borderwidth=0, highlightthickness=2, highlightbackground="grey", highlightcolor="grey")
        self.note.grid(column=2, row=0, columnspan=6,
                       padx=(4, 4), pady=0, sticky="WE")

        self.openBtn = ttk.Button(
            self, text="open", command=self.openExcelFile, style="my.TButton")
        self.openBtn.place(relwidth=0.3, relheight=0.3, x=4, y=55)

        self.exportBtn = ttk.Button(
            self, text="export", command=self.export2xlsx, style="my.TButton")
        self.exportBtn.place(relwidth=0.3, relheight=0.3, x=120, y=55)

    def export2xlsx(self):
        wb = Workbook()
        if self.filename != "":
            wb = load_workbook(self.filename)
        else:
            wb.create_sheet("all")
            wb.create_sheet("once")

        ws = wb["all"]
        ws.append(["note", self.note.get()])
        ws.append(self.master.master.dataName)

        for i in range(len(self.master.master.data[0])-1):
            ws.append(self.master.master.data[j][i+1]
                      for j in range(len(self.master.master.data)))

        ws.append([])
        if self.filename != "":
            wb.save(self.filename)
        else:
            wb.save(
                "./excel/"+datetime.datetime.now().strftime("%d-%m-%y %H:%M")+".xlsx")

    def openExcelFile(self):
        self.filename = filedialog.askopenfilename(
            initialdir="./excel", title="Select a excel file", filetypes=(("all", "."), ("xlsx", "*.xlsx")))

    # def autoExport(self):
    #     currentTime = datetime.datetime.now()
    #     packetArray = recentPacketString.strip("\n").split("-")
    #     if packetArray[0] == "ae":
    #         print("found ae")
    #         packetArray.pop(0)
    #         print(packetArray)
    #         ws_auto.append(packetArray)
    #     if packetArray[0] == "end":
    #         print("found end")
    #         ws_auto.append([])
    #         wb_auto.save("./excel/"+currentTime.strftime("%d-%m-%y")+"_auto.xlsx")
